from openpyxl import load_workbook
import requests
import threading
import datetime
from app import app

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def read_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

def post_to_vestaboard(message):
    response = requests.post(
        "https://vbml.vestaboard.com/format",
        headers={"Content-Type": "application/json"},
        json={"message": message}
    )
    if response.status_code == 200:
        print(f"Successfully updated Vestaboard with:\n{message}")
    else:
        print(f"Failed to update Vestaboard: {response.status_code}")

def schedule_trips(trips):
    trips_by_time = {}
    for trip in trips:
        departure_time = trip['Departure']
        if departure_time not in trips_by_time:
            trips_by_time[departure_time] = []
        trips_by_time[departure_time].append(trip)

    for departure_time, trips_at_time in trips_by_time.items():
        departure_datetime = datetime.datetime.strptime(str(departure_time), '%H:%M:%S').time()
        now = datetime.datetime.now()
        departure = datetime.datetime.combine(now.date(), departure_datetime)
        
        delay = (departure - datetime.datetime.now()).total_seconds()
        
        if delay > 0:
            message = "\n".join([f"{trip['Customer']}\n{trip['From']} -> {trip['To']}" for trip in trips_at_time])
            threading.Timer(delay - 3600, post_to_vestaboard, args=(message,)).start()
            print(f"Scheduled batch for {len(trips_at_time)} trips at {departure_time}")
        else:
            print(f"Trips at {departure_time} already passed today.")