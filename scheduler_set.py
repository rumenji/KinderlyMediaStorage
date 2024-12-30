from openpyxl import load_workbook
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.jobstores.memory import MemoryJobStore
from apscheduler.executors.pool import ThreadPoolExecutor, ProcessPoolExecutor
import requests
import datetime
import os
from dotenv import load_dotenv
from app import app
from helpers import get_last_name, replace_sides

load_dotenv()

# Load predifined environment variables
# Vestaboard subscription ID, key, and secret - obtained from Vetaboard
SUBSCRIPTION_ID = os.getenv("SUBSCRIPTION_ID")
X_VESTABOARD_API_KEY = os.getenv("X-VESTABOARD-API-KEY")
X_VESTABOARD_API_SECRET = os.getenv("X-VESTABOARD-API-SECRET")
# Default message to show when no trips are currently scheduled, and color for the border
DEFAULT_MESSAGE = os.getenv('DEFAULT_MESSAGE')
DEFAULT_MESSAGE_COLOR = os.getenv('DEFAULT_MESSAGE_COLOR')
# Title to display on first row - e.g "Next departures:"
TITLE = os.getenv('TITLE')
# Value to filter rows from the uploaded spreadhseet.
SEARCH_VALUE = os.getenv('SEARCH_VALUE')

# Create the APScheduler Job store. Will try to repeat a misfired job for 90 minutes after initial scheduled run
job_store = MemoryJobStore()
executors = {'default': ThreadPoolExecutor(20),
             'processpool': ProcessPoolExecutor(5)}
job_defaults = {
    'coalesce': False,
    'max_instances': 3,
    'misfire_grace_time': 90
}
scheduler = BackgroundScheduler(jobstores={'default': job_store}, executors=executors, job_defaults=job_defaults)

def read_excel(file_path):
    '''Read the uploaded spreadsheet. Checks is the required columns are in the spreadsheet.
    Checks if the spreadsheet contains any trips.
    Only returns rows for Retail customers, and if the destination SEARCH_VALUE matches'''
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        required_columns = ['Start time LT', 'End time LT', 'Customer', 'Orig', 'Dest']
        if not all(col in headers for col in required_columns):
            raise Exception('The file does not have columns Start time LT, End time LT, Customer, Orig, or Dest')
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        data = [trip for trip in data if trip['Customer type'] == 'Retail' and any([trip['Orig'] == SEARCH_VALUE, trip['Dest'] == SEARCH_VALUE])]
        if len(data) == 0:
            raise Exception('No matching trips in the uploaded file!')
        return data
    except Exception as e:
        raise e

# Keeping the currently displayed trips in memory to send to Vestaboard
# Vestaboard cannot add - only replace what is displayed.
CURRENTLY_DISPLAYED_TRIPS = []

def format_vestaboard_message(message, time, last):
    try:
        global CURRENTLY_DISPLAYED_TRIPS
        if last:
            CURRENTLY_DISPLAYED_TRIPS = []
            formatted_msg = requests.post(
                "https://vbml.vestaboard.com/compose",
                headers={"Content-Type": "application/json"},
                json={"components": [{
			                    "style": {
				                "justify": "center",
                                "align": "center"
                                },
                                "template": message
                                }]}
                )
            
            if formatted_msg.status_code == 200:
                return replace_sides(formatted_msg.json(), DEFAULT_MESSAGE_COLOR)
            else:
                raise Exception(f"Failed to format message: {formatted_msg.status_code}")
        else:
            # If the current trips list is empty - add the title
            if len(CURRENTLY_DISPLAYED_TRIPS) == 0:
                CURRENTLY_DISPLAYED_TRIPS.append({
                    "msg": {
                    "style": {
                    "height": 1,
                    "width": 22
                    },
                    "template": TITLE
                },
                "time": datetime.datetime.strptime('11:59 pm', '%I:%M %p').time()
                })
            # If it is not empty - filter the trips by time and remove the ones that have already left 15 minutes ago.
            else:
                now = datetime.datetime.now()
                upcoming_trips = [trip for trip in CURRENTLY_DISPLAYED_TRIPS if trip['time'] > now.time() - datetime.timedelta(minutes=15)]
                CURRENTLY_DISPLAYED_TRIPS = upcoming_trips
            # Add the current trip to list
            CURRENTLY_DISPLAYED_TRIPS.append(
                    {
                        "msg": {
                        "style": {
                        "height": 1,
                        "width": 22,
                        "align": "center"
                        },
                        "template": message
                    },
                    "time": time
                    }
                )
            # Send the request to the Vestaboard format API
            formatted_msg = requests.post(
                "https://vbml.vestaboard.com/compose",
                headers={"Content-Type": "application/json"},
                json={"components": [trip['msg'] for trip in CURRENTLY_DISPLAYED_TRIPS]}
            )
            # Return the formatted list
            if formatted_msg.status_code == 200:
                print(CURRENTLY_DISPLAYED_TRIPS)
                return formatted_msg.json()
            else:
                raise Exception(f"Failed to format message: {formatted_msg.status_code}")
    except Exception as e:
        raise e
    
    
def post_to_vestaboard(message, time, last):
    '''Send the request to update the board.
    Expects the message to display, time until it should be displayed, and if it is the last message -
    to display the default message with proper formatting'''
    try:
        # Format the message using the Vestaboard API
        characters = format_vestaboard_message(message, time, last)
        # If the format is successful, send the request to update the bord
        if characters:
            response = requests.post(
                f"https://subscriptions.vestaboard.com/subscriptions/{SUBSCRIPTION_ID}/message",
                headers={"Content-Type": "application/json",
                        "x-vestaboard-api-key": X_VESTABOARD_API_KEY,
                        "x-vestaboard-api-secret": X_VESTABOARD_API_SECRET},
                json={"characters": characters}
            )
            if response.status_code != 200:
                raise Exception(f"Failed to update Vestaboard: {response.status_code}")
    except Exception as e:
        raise e

def schedule_trips(trips):
    '''Schedule jobs in the store. Expects a list of all trips'''
    if len(trips) == 0:
        raise Exception('No trips in the uploaded file!')
    # A variable to set the last departure\arival time. Used to set when the default message should be displayed
    last_departure = datetime.datetime.now().time()
    # A list to keep all skipped trips - trips that have already passed.
    skipped_trips = []
    # Iterate over the trips and schedule the jobs for each.
    for trip in trips:
        file_departure = trip.get('Start time LT')
        file_arrival = trip.get('End time LT')
        file_customer = trip.get('Customer')
        file_pax = trip.get('Pax allocated')
        file_origin = trip.get('Orig')
        file_destination = trip.get('Dest')
        # Check if all values from the columns exist
        if not all([file_departure, file_arrival, file_customer, file_destination, file_origin]):
            raise Exception('The file does not have columns Start time LT, End time LT, Customer, Orig, or Dest')
        # If departure - show departure time, if arrival - show arrival time
        time_to_show = file_departure if file_origin == SEARCH_VALUE else file_arrival
        # Only schedule a trip if there are passengers
        if file_pax != 0:
            try:
                # Get the last name of the passenger using the helper function - limit to 9 characters 
                customer_name = get_last_name(file_customer)
                # Convert the departure/arrival time from string in the format HH:MM AM/PM to datetime object time only.
                departure_time = datetime.datetime.strptime(str(time_to_show), '%I:%M %p').time()
                # Get the current time to make sure only future trips are scheduled
                now = datetime.datetime.now()
                # Add date to the time of the trips
                departure = datetime.datetime.combine(now.date(), departure_time)
                # Calculate if the date and time of the trip are in the future
                delay = (departure - now).total_seconds()
                if delay > 0:
                    # Run the job one hour before the trip, or if it is still in the future but less than one hour away - run it right away
                    job_runtime = departure - datetime.timedelta(hours=1) if delay>3600 else now
                    # Schedue the job. The args are msg - customer name + orig-dest + time HHMM, time - to stay on the board until, last message - bool 
                    job = scheduler.add_job(
                        post_to_vestaboard, 
                        'date', 
                        run_date=job_runtime,
                        args=[f"{customer_name} {file_origin[-3:]}-{file_destination[-3:]} {departure_time.strftime('%H%M')}", departure_time, False],
                        id=f"{file_customer}_{file_origin}_{file_destination}"
                    )
                    # Update last departure if the trip is after the current value
                    if last_departure < departure_time:
                        last_departure = departure_time
                # Add skipped if trips in the file were in the past to display a message for the user
                else:
                    skipped_trips.append(f"{file_customer} at {departure_time}")
            except Exception as e:
                raise e
    #Append the end default message to display 30 minutes after the last departure/arrival
    try:  
        scheduler.add_job(
                    post_to_vestaboard, 
                    'date', 
                    run_date=datetime.datetime.combine(now.date(), last_departure) + datetime.timedelta(minutes=30),
                    args=[DEFAULT_MESSAGE, last_departure, True],
                    id="End"
                )
    except Exception as e:
        raise e
    # Return the skipped trips list to the route
    return skipped_trips

scheduler.start()  # Start the scheduler